from pyscipopt import Model, quicksum
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from data_processing import main







# Statt MA könnte man hier User_id verwenden
# 40 MA, 10h
MA1 = [1, 1, 1, 1, 1, 1, 0, 0, 0, 0]
MA2 = [0, 1, 1, 1, 1, 1, 1, 0, 0, 0]
MA3 = [0, 0, 1, 1, 1, 1, 1, 1, 0, 0]
MA4 = [0, 0, 0, 1, 1, 1, 1, 1, 1, 0]
MA5 = [0, 0, 0, 0, 1, 1, 1, 1, 1, 1]
MA6 = [1, 1, 1, 1, 1, 1, 0, 0, 0, 0]
MA7 = [0, 1, 1, 1, 1, 1, 1, 0, 0, 0]
MA8 = [0, 0, 1, 1, 1, 1, 1, 1, 0, 0]
MA9 = [0, 0, 0, 1, 1, 1, 1, 1, 1, 0]
MA10 = [0, 0, 0, 0, 1, 1, 1, 1, 1, 1]
MA11 = [1, 1, 1, 1, 1, 1, 0, 0, 0, 0]
MA12 = [0, 1, 1, 1, 1, 1, 1, 0, 0, 0]
MA13 = [0, 0, 1, 1, 1, 1, 1, 1, 0, 0]
MA14 = [0, 0, 0, 1, 1, 1, 1, 1, 1, 0]
MA15 = [0, 0, 0, 0, 1, 1, 1, 1, 1, 1]
MA16 = [1, 1, 1, 1, 1, 1, 0, 0, 0, 0]
MA17 = [0, 1, 1, 1, 1, 1, 1, 0, 0, 0]
MA18 = [0, 0, 1, 1, 1, 1, 1, 1, 0, 0]
MA19 = [0, 0, 0, 1, 1, 1, 1, 1, 1, 0]
MA20 = [0, 0, 0, 0, 1, 1, 1, 1, 1, 1]
MA21 = [0, 0, 0, 0, 1, 1, 1, 1, 1, 1]
MA22 = [0, 1, 1, 1, 1, 1, 1, 0, 0, 0]
MA23 = [0, 0, 1, 1, 1, 1, 1, 1, 0, 0]
MA24 = [0, 0, 0, 1, 1, 1, 1, 1, 1, 0]
MA25 = [0, 0, 0, 0, 1, 1, 1, 1, 1, 1]
MA26 = [1, 1, 1, 1, 1, 1, 0, 0, 0, 0]
MA27 = [0, 1, 1, 1, 1, 1, 1, 0, 0, 0]
MA28 = [0, 0, 1, 1, 1, 1, 1, 1, 0, 0]
MA29 = [0, 0, 0, 1, 1, 1, 1, 1, 1, 0]
MA30 = [0, 0, 0, 0, 1, 1, 1, 1, 1, 1]
MA31 = [0, 0, 0, 0, 1, 1, 1, 1, 1, 1]
MA32 = [0, 1, 1, 1, 1, 1, 1, 0, 0, 0]
MA33 = [0, 0, 1, 1, 1, 1, 1, 1, 0, 0]
MA34 = [0, 0, 0, 1, 1, 1, 1, 1, 1, 0]
MA35 = [0, 0, 0, 0, 1, 1, 1, 1, 1, 1]
MA36 = [1, 1, 1, 1, 1, 1, 0, 0, 0, 0]
MA37 = [0, 1, 1, 1, 1, 1, 1, 0, 0, 0]
MA38 = [0, 0, 1, 1, 1, 1, 1, 1, 0, 0]
MA39 = [0, 0, 0, 1, 1, 1, 1, 1, 1, 0]
MA40 = [0, 0, 0, 0, 1, 1, 1, 1, 1, 1]

availability = [MA1, MA2, MA3, MA4, MA5, MA6, MA7, MA8, MA9, MA10,
                MA11, MA12, MA13, MA14, MA15, MA16, MA17, MA18, MA19, MA20,
                MA21, MA22, MA23, MA24, MA25, MA26, MA27, MA28, MA29, MA30,
                MA31, MA32, MA33, MA34, MA35, MA36, MA37, MA38, MA39, MA40]

# Mitarbeiterkosten
employee_costs = [100, 100, 100, 100, 100,
                  15, 15, 15, 15, 15,
                  20, 20, 20, 20, 20,
                  25, 25, 25, 25, 25,
                  30, 30, 30, 30, 30,
                  35, 35, 35, 35, 35,
                  40, 40, 40, 40, 40,
                  5, 5, 5, 5, 5]

# Max. tägliche Arbeitsstunden für jeden Mitarbeiter
max_daily_hours = [5, 5, 5, 5, 5,
                  5, 5, 5, 5, 5,
                  5, 5, 5, 5, 5,
                  5, 5, 5, 5, 5,
                  5, 5, 5, 5, 5,
                  5, 5, 5, 5, 5,
                  5, 5, 5, 5, 5,
                  1, 1, 10, 1, 1]


# Benötigte Mitarbeiter pro Stunde
required_employees = [5, 5, 10, 10, 1, 25, 20, 10, 5, 2]

"""
# Friendnumber für jeden Mitarbeiter
# --> zusätzlich noch einen Faktor hinzufügen, ob friendnumber oder Kostenminimierung mehr gewichtet werden soll
friend_numbers = [1, 0, 0, 0, 0, 0, 0, 0, 0, 0,
                 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
                 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
                 0, 0, 0, 0, 0, 0, 0, 0, 0, 1]
"""

# Anzahl Mitarbeiter
num_employees = len(availability)

# Anzahl Stunden
num_hours = len(availability[0])

# Die max. Arbeitsblöcke pro Tag pro MA einstellen
max_work_blocks = [1] * num_employees


# Optimierungsmodell erstellen -----------------------------------------------------------------------------------------

model = Model("40 MA, 10h")

# Entscheidungsvariable x (MA arbeitet = 1, nicht arbeiten = 0 / für NB 1 + 2)
x = {}
for i in range(num_employees):
    for t in range(num_hours):
        x[i, t] = model.addVar(f"x_{i}_{t}", vtype="B")

print()
print("Entscheidungsvariable x: ", x)

# Entscheidungsvariablen y (neuer Arbeitsblock beginnen = 1, nicht = 0 / für NB 5 + 6)
y = {}
for i in range(num_employees):
    for t in range(num_hours):
        y[i, t] = model.addVar(f"y_{i}_{t}", vtype="B")

print("Entscheidungsvariable y: ", y)
print()

# Für ganzzahlige Variablen (Vom Wert 0 - 10)
# x[i, t] = model.addVar(f"x_{i}_{t}", vtype="I", lb=0, ub=10)

# Für stetige Variablen
# x[i, t] = model.addVar(f"x_{i}_{t}", vtype="C", lb= tiefste_zahl, ub= höchste-zahl)


# Nebenbedingungen -----------------------------------------------------------------------------------------------------

# NB1: Anzahl der benötigten Mitarbeiter pro Stunde
for t in range(num_hours):
    temp_sum = 0
    for i in range(num_employees):
        temp_sum += x[i, t]
    model.addCons(temp_sum == required_employees[t])

# NB2: jeder MA arbeitet mind. 1h
for i in range(num_employees):
    model.addCons(quicksum(x[i, t] for t in range(num_hours)) >= 1)

# NB3: Verfügbarkeitsbeschränkung
for i in range(num_employees):
    for t in range(num_hours):
        model.addCons(x[i, t] <= availability[i][t])

# NB4: max. täglichen Arbeitsstunden dürfen nicht überschritten werden
for i in range(num_employees):
    model.addCons(quicksum(x[i, t] for t in range(num_hours)) <= max_daily_hours[i])


# ---------------------------------------  NB5 - 7 nochmal überarbeiten  -----------------------------------------------

# NB5: Wenn ein Mitarbeiter in einer Stunde arbeitet, muss er in den folgenden Stunden entweder weiterarbeiten
# oder nicht mehr arbeiten.
for i in range(num_employees):
    for t in range(num_hours - 1):
        model.addCons(y[i, t] >= x[i, t] - x[i, t + 1])
        model.addCons(y[i, t] <= x[i, t])

# NB6: Maximale Anzahl der Arbeitsblöcke für jeden Mitarbeiter
for i in range(num_employees):
    model.addCons(quicksum(y[i, t] for t in range(num_hours)) <= max_work_blocks[i])

# NB7: Nur einen Arbeitsblock pro Mitarbeiter erlauben
for i in range(num_employees):
    for t in range(1, num_hours - 1):
        model.addCons(x[i, t - 1] + x[i, t + 1] <= 1 + x[i, t])


# Hauptbedingung -------------------------------------------------------------------------------------------------------

model.setObjective(quicksum(employee_costs[i] * x[i, t] for i in range(num_employees) for t in range(num_hours)),
                       "minimize")

# Modell lösen ---------------------------------------------------------------------------------------------------------

model.optimize()

# Überprüfung und Ausgabe ----------------------------------------------------------------------------------------------
print()
if model.getStatus() == "optimal":
    schedule = [[int(model.getVal(x[i, t])) for t in range(num_hours)] for i in range(num_employees)]
    print("Zeitplan optimiert:")
    for i, row in enumerate(schedule):
        print(f"MA{i + 1}: {row}")
else:
    print("Keine optimale Lösung gefunden")

# Eingegebene / eingesetzte Stunden kompl. -----------------------------------------------------------------------------

eingegebene_stunden = 0
for i in range(num_employees - 1):
    for j in range(num_hours - 1):
        if availability[i][j] == 1:
            eingegebene_stunden += 1

print()
print("Anzahl eingegebener Stunden: ", eingegebene_stunden)

eingesetzte_stunden = 0
for e in required_employees:
    eingesetzte_stunden += e

print("Anzahl eingesetzter Stunden: ", eingesetzte_stunden)

# in eine Excel-Datei exportieren --------------------------------------------------------------------------------------

# Excel-Datei und Arbeitsblatt erstellen
workbook = Workbook()
worksheet = workbook.active

# Definieren Sie die Farben für die Zellen
no_work_fill = PatternFill(start_color="FFC0C0", end_color="FFC0C0", fill_type="solid")
work_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

# Zellen des Arbeitsblatts basierend auf der berechneten Matrix füllen
for i, row in enumerate(schedule):
    for j, cell_value in enumerate(row):
        cell = worksheet.cell(row=i + 2, column=j + 2)  # Zeilen und Spalten um 1 erhöht
        cell.value = cell_value
        cell.fill = work_fill if cell_value == 1 else no_work_fill

# Spalten- und Zeilenüberschriften hinzufügen
for j in range(num_hours):
    worksheet.cell(row=1, column=j + 2).value = f"Stunde {j + 1}"
for i in range(num_employees):
    worksheet.cell(row=i + 2, column=1).value = f"MA{i + 1}"

# Excel-Datei speichern
workbook.save("Einsatzplan.xlsx")

"""
solutions = model.getSols()
for i, solution in enumerate(solutions):
    print(f"Lösung {i + 1}:")
    model.printSol(solution)
"""